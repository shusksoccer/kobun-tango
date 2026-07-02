# -*- coding: utf-8 -*-
import pandas as pd, re, json, os
ROOT=os.path.dirname(os.path.abspath(__file__))

SRC = os.path.join(ROOT,'data','kobun-words.xlsx')
df = pd.read_excel(SRC, sheet_name='リストa', header=1)
df.columns = ['idxA','idxB','no','word','pos','keigo','imi','ko','koBlank','yk','ykBlank',
              'no2','word2','imi2','ko2','koBlank2','yk2','ykBlank2']
df = df.dropna(subset=['no']).copy(); df['no']=df['no'].astype(int)

cl_imi  = lambda s: re.sub(r'[〔〕\s　]','',str(s)).strip() if pd.notna(s) else ''
cl_txt  = lambda s: str(s).strip() if pd.notna(s) else ''
def cl_yk(s):
    s=cl_txt(s); return re.sub(r'^[　\s]*[（(]訳[）)]','',s).strip()
def src_of(ko):
    m=re.findall(r'（([^（）]+)）',cl_txt(ko)); return m[-1] if m else ''
def strip_src(ko):
    return re.sub(r'\s*（[^（）]+）\s*$','',cl_txt(ko)).strip()

BL_RE=re.compile(r'〔[\s　]*[〕\]）]')
def ko_underline(ko, koBlank):
    s=strip_src(ko); kb=strip_src(koBlank)
    if '〔' not in kb: return s
    lits=BL_RE.split(kb)
    if len(lits)<2: return s
    rest=s; out=''; ok=True
    for i in range(len(lits)-1):
        lit=lits[i]; nxt=lits[i+1]
        if lit:
            p=rest.find(lit)
            if p<0: ok=False; break
            out+=rest[:p+len(lit)]; rest=rest[p+len(lit):]
        q=rest.find(nxt) if nxt else len(rest)
        if q<0: ok=False; break
        out+='\x01'+rest[:q]+'\x02'; rest=rest[q:]
    out+=rest
    return out if ok else s

# ---------- meaning consolidation (merge example-answers of the same sense) ----------
try:
    import fugashi
    _TAGGER=fugashi.Tagger()
    def lemmatize(label):
        try:
            ws=_TAGGER(label)
            if not ws: return label
            last=ws[-1]
            base=getattr(last.feature,'orthBase',None) or getattr(last.feature,'lemma',None) or last.surface
            return label[:len(label)-len(last.surface)]+base
        except Exception:
            return label
except Exception:
    def lemmatize(label): return label
ALLOWED={'', 'る','れ','れる','られ','られる','り','ら','ろ','っ','た','て','だ','で',
 'い','く','き','かっ','かろ','けれ','しい','しく','しき','しかっ','じ',
 'こと','ること','の','ものだ',
 'な','に','なり','なる','なら','なっ','なれ','だっ','であ','である',
 'ず','ぬ','ね','ない','なく','なかっ',
 'する','すれ','しろ','せよ','した','して',
 'う','よ','よう','ます','まし','ませ','ましょ','さ','せ','し','す','む','み','ま','め','ば','べ','わ','を','ん','も','や'}
TERM=('る','い','な','り','だ','なり','う','む','ぐ','ぶ','つ','ぬ','す','ない','する')
OVERRIDE={'亡くなっ':'亡くなる','いらっしゃっ':'いらっしゃる','召し上がら':'召し上がる',
 'お召しになっ':'お召しになる','不平を言った':'不平を言う','管絃を楽しみ':'管絃を楽しむ',
 'お召しになれ':'お召しになる','妨げられず':'妨げられる','華やかに':'華やかだ','突然の':'突然だ'}
def _collapse(s):
    n=len(s)
    return s[:n//2] if (n>=2 and n%2==0 and s[:n//2]==s[n//2:]) else s
def _cp(a,b):
    i=0
    while i<len(a) and i<len(b) and a[i]==b[i]: i+=1
    return i
def _same(a,b):
    a=_collapse(a); b=_collapse(b)
    if a==b: return True
    c=_cp(a,b)
    if c<1: return False
    return a[c:] in ALLOWED and b[c:] in ALLOWED
def cluster_meanings(ms):
    cs=[]
    for m in ms:
        for c in cs:
            if any(_same(m,x) for x in c): c.append(m); break
        else: cs.append([m])
    return cs
def canon_label(c):
    def sc(s):
        s2=_collapse(s); return (1 if s2.endswith(TERM) else 0, len(s2))
    lab=_collapse(max(c,key=sc))
    return OVERRIDE.get(lab,lab)

def importance(no):
    if no<=50:return '最必修'
    if no<=150:return '必修'
    if no<=180:return '必修敬語'
    if no<=280:return '重要'
    return '応用'

KEIGO={151:['尊敬'],152:['尊敬'],153:['尊敬'],154:['尊敬'],155:['謙譲'],156:['謙譲'],
 157:['尊敬'],158:['尊敬'],159:['尊敬'],160:['丁寧'],161:['尊敬'],162:['謙譲'],163:['謙譲'],
 164:['謙譲','尊敬'],165:['丁寧','謙譲'],166:['丁寧','謙譲'],167:['尊敬'],168:['謙譲'],
 169:['謙譲','尊敬'],170:['謙譲'],171:['謙譲'],172:['謙譲'],173:['尊敬'],174:['尊敬'],
 175:['謙譲'],176:['尊敬'],177:['尊敬'],178:['尊敬'],179:['謙譲'],180:['謙譲']}

# 意味テーマ（docs/TAG_GUIDE.md が定義・判断基準の正）。
# 領域8種＋評価2種の10種、多ラベル（該当すれば重複付与）。中立・機能語は無印。
TL={'LOVE':'恋・男女','DEATH':'病・死・別れ','BUD':'仏道・信仰','COURT':'宮廷・身分',
 'TASTE':'情趣','TIME':'時・時間','WORD':'ことば・学問','THINK':'思考・評判',
 'PLUS':'プラスの評価・心情','MINUS':'マイナスの評価・心情'}
THEME_ORDER=['LOVE','DEATH','BUD','COURT','TASTE','TIME','WORD','THINK','PLUS','MINUS']
TH={1:['THINK'],2:['THINK'],3:['BUD'],4:['THINK'],5:['LOVE'],6:['MINUS'],7:['LOVE','THINK'],8:['LOVE'],
 11:['PLUS'],12:['PLUS'],13:['PLUS'],14:['PLUS','MINUS'],15:['TASTE','PLUS'],16:['PLUS','COURT'],
 17:['MINUS','COURT'],18:['PLUS'],19:['THINK','PLUS'],20:['MINUS'],21:['PLUS'],22:['PLUS'],23:['MINUS'],
 24:['MINUS'],25:['MINUS'],26:['TASTE','PLUS'],27:['MINUS'],28:['PLUS','LOVE'],29:['MINUS'],30:['MINUS'],
 31:['TIME'],34:['BUD'],35:['WORD'],36:['WORD'],38:['TASTE','LOVE','PLUS'],39:['LOVE','BUD'],45:['TIME'],
 51:['PLUS'],52:['LOVE'],54:['TASTE'],55:['PLUS'],60:['THINK'],66:['DEATH','BUD'],69:['DEATH','MINUS'],
 70:['MINUS'],72:['DEATH'],73:['BUD'],74:['TASTE','PLUS'],75:['PLUS'],76:['COURT','PLUS'],
 77:['PLUS','MINUS'],78:['PLUS'],79:['PLUS'],80:['MINUS'],81:['PLUS'],82:['PLUS'],84:['MINUS'],
 85:['MINUS'],86:['PLUS','MINUS'],87:['MINUS'],88:['PLUS'],89:['LOVE','MINUS'],90:['MINUS'],91:['MINUS'],
 92:['MINUS'],93:['MINUS'],94:['MINUS'],95:['MINUS'],96:['MINUS'],97:['MINUS'],98:['MINUS'],
 99:['COURT','PLUS'],100:['PLUS'],103:['TIME'],104:['MINUS'],105:['LOVE','MINUS'],107:['MINUS'],
 108:['COURT'],109:['COURT'],110:['BUD'],113:['BUD'],114:['THINK','LOVE'],115:['DEATH'],117:['TIME'],
 118:['TIME'],121:['LOVE'],124:['WORD'],125:['WORD'],127:['TIME'],128:['TIME'],129:['TIME'],145:['TIME'],
 147:['THINK'],149:['PLUS'],174:['TASTE'],178:['COURT'],179:['COURT'],180:['COURT'],181:['COURT'],
 188:['LOVE'],189:['LOVE'],193:['PLUS'],195:['LOVE','COURT'],197:['DEATH'],198:['MINUS'],199:['MINUS'],
 200:['MINUS'],201:['DEATH'],202:['DEATH'],203:['DEATH'],204:['DEATH'],205:['DEATH'],207:['MINUS'],
 208:['PLUS'],209:['MINUS'],210:['PLUS'],211:['PLUS','MINUS'],212:['LOVE','TASTE'],213:['PLUS'],
 214:['PLUS'],215:['PLUS','MINUS'],216:['PLUS'],217:['PLUS'],218:['PLUS'],219:['PLUS'],220:['MINUS'],
 221:['MINUS'],222:['MINUS'],223:['PLUS'],224:['MINUS'],225:['MINUS'],226:['PLUS','MINUS'],
 227:['PLUS','MINUS'],228:['MINUS'],229:['MINUS'],230:['MINUS'],231:['TIME','MINUS'],232:['MINUS'],
 233:['MINUS'],234:['MINUS'],235:['MINUS'],236:['PLUS'],237:['MINUS'],238:['MINUS'],240:['DEATH'],
 241:['MINUS'],242:['MINUS'],243:['MINUS'],244:['PLUS','MINUS'],245:['MINUS'],246:['PLUS'],249:['MINUS'],
 250:['TIME'],251:['LOVE'],252:['PLUS'],253:['PLUS'],254:['PLUS'],255:['COURT'],256:['COURT'],
 257:['COURT'],259:['TASTE','PLUS'],261:['WORD'],263:['WORD'],264:['WORD'],265:['COURT'],267:['LOVE'],
 272:['MINUS'],273:['TIME'],275:['TIME'],280:['MINUS'],281:['THINK'],285:['MINUS'],286:['MINUS'],
 287:['COURT'],288:['PLUS'],289:['MINUS'],290:['MINUS'],291:['MINUS'],293:['PLUS','MINUS'],294:['MINUS'],
 295:['PLUS'],298:['LOVE','TASTE'],299:['TASTE'],300:['TASTE'],309:['TIME'],310:['COURT'],311:['TIME'],
 312:['TIME'],322:['MINUS'],323:['MINUS'],325:['MINUS'],326:['LOVE'],329:['TIME']}

# 活用情報（動詞＝行・活用型、形容詞＝ク/シク）。形容動詞は一律ナリ活用（付与時に自動）。
# 四段/下二段の両用で意味が変わる語（58たのむ・71かづく）は両方を明記。
KATSUYO={
 1:'カ行四段',2:'ラ行四段',3:'サ変',4:'ヤ行下二段',5:'バ行上二段',6:'マ行下二段',7:'ヤ行下二段',
 8:'ハ行四段',9:'ワ行上一段',10:'カ行四段',11:'カ行四段',
 51:'ハ行四段',52:'ハ行四段',53:'ナ変',54:'バ行四段',55:'ダ行下二段',56:'ハ行四段',57:'カ行四段',
 58:'マ行四段（あてにする）／下二段（あてにさせる）',59:'ラ行四段',60:'ヤ行下二段',61:'サ行下二段',
 62:'カ行下二段',63:'ハ行下二段',64:'ラ行四段',65:'サ変',66:'ハ行四段',67:'サ変',68:'カ行四段',
 69:'ハ行四段',70:'ハ行四段',71:'カ行四段（いただく）／下二段（与える）',72:'サ行下二段',73:'ハ行四段',
 151:'ハ行四段／サ行下二段',152:'サ行下二段',153:'サ行四段',154:'サ変',155:'サ行四段',
 156:'ヤ行下二段／サ行下二段',157:'サ変／サ行四段',158:'サ行四段',159:'ハ行四段',160:'ハ行下二段',
 161:'サ行下二段',162:'ラ行四段',163:'ラ行四段',164:'ラ行四段',165:'ラ変',166:'ハ行四段',167:'サ行四段',
 168:'ダ行下二段',169:'ラ行四段',170:'サ行下二段',171:'ラ行四段',172:'ダ行下二段',173:'サ行四段',
 174:'サ行四段',175:'ラ行四段',176:'ラ行四段',177:'サ行四段',179:'サ変',180:'サ変',
 181:'ラ行四段',182:'ラ行下二段',183:'ハ行四段',184:'ダ行下二段',185:'サ行四段',186:'ハ行四段',
 187:'マ行四段',188:'ハ行四段',189:'マ行四段',190:'ラ行四段',191:'サ変',192:'サ行四段',193:'サ変',
 194:'ガ行下二段',195:'カ行四段',196:'ラ行四段',197:'サ変',198:'サ行四段',199:'ハ行下二段',
 200:'タ行四段',201:'ラ行四段',202:'ハ行四段',203:'ラ行四段',204:'マ行四段',205:'ラ行下二段',
 206:'マ行下二段',281:'ラ行四段',282:'タ行下二段',283:'ハ行四段',284:'サ行四段',285:'ラ行下二段',
 286:'ラ行四段',287:'タ行四段',
 12:'シク活用',13:'シク活用',14:'シク活用',15:'シク活用',16:'ク活用',17:'シク活用',18:'シク活用',
 19:'シク活用',20:'ク活用',21:'ク活用',22:'ク活用',23:'シク活用',24:'ク活用',25:'シク活用',
 74:'ク活用',75:'シク活用',76:'ク活用',77:'シク活用',78:'ク活用',79:'シク活用',80:'ク活用',
 81:'シク活用',82:'シク活用',83:'ク活用',84:'ク活用',85:'ク活用',86:'シク活用',87:'シク活用',
 88:'ク活用',89:'ク活用',90:'ク活用',91:'ク活用',92:'ク活用',93:'シク活用',94:'ク活用',95:'ク活用',
 96:'ク活用',97:'シク活用',98:'シク活用',
 207:'シク活用',208:'ク活用',209:'シク活用',210:'ク活用',211:'シク活用',212:'シク活用',213:'シク活用',
 214:'シク活用',215:'ク活用',216:'シク活用',217:'シク活用',218:'ク活用',219:'ク活用',220:'ク活用',
 221:'シク活用',222:'ク活用',223:'シク活用',224:'ク活用',225:'ク活用',226:'ク活用',227:'ク活用',
 228:'シク活用',229:'ク活用',230:'ク活用',231:'シク活用',232:'ク活用',233:'ク活用',234:'ク活用',
 235:'ク活用',236:'ク活用',237:'シク活用',238:'ク活用',239:'ク活用',240:'シク活用',
 288:'シク活用',289:'ク活用',290:'ク活用',291:'ク活用',292:'ク活用',293:'ク活用',294:'シク活用',295:'シク活用'}

# learning-flag sets（docs/TAG_GUIDE.md が定義・判断基準の正）
GAP_BIG={2,15,17,19,21,26,38,45,59,66,73,77,82,84,86,93,98,99,103,105,114,137,140,147,181,182,183,
 195,202,203,205,207,208,237,240,314}
GAP_SMALL={1,5,7,8,9,11,12,13,16,18,25,28,55,58,60,69,74,78,79,87,89,119,141,190,204,210,212,213,217,
 222,223,230,258,260,271,274,283,295,298,301,305,326}
CORE={1,4,5,7,10,14,15,20,26,37,38,40,60,64,67,77,78,85,120,233}
VAGUE={10,15,26,37,40,64,67,106,120,122,233,260,268,300,302,303}  # 漠然系：意味の輪郭がぼやける語（TAG_GUIDE §3-(1)(2)）
KANGO={3,65,179,180,191,193,197}
KEIGO_CARE={153,156,159,160,164,165,166,167,169}
KOOU_EXTRA={50,143,144,318}  # 「～」を含まないが呼応の副詞に該当する語

# ---------- authoritative meanings (意味一覧シート) ----------
# 各語の意味・意味数は data/古文単語_意味付き.xlsx の「意味一覧」シートを正とする。
# build 内の自動抽出（活用違いのクラスタリング）は例文↔意味の対応（mi）を出すためだけに使い、
# 表示する意味リストはここで上書きする。
import openpyxl, difflib
MEAN_SRC=os.path.join(ROOT,'data','古文単語_意味付き.xlsx')
def load_auth_meanings(path):
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws=next((s for s in wb.worksheets if '意味一覧' in s.title), wb.worksheets[0])
    d={}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        ms=[str(c).strip() for c in row[3:9] if c is not None and str(c).strip()]
        if ms: d[int(row[0])]=ms
    return d
AUTH=load_auth_meanings(MEAN_SRC)
def _mnorm(s):
    s=re.sub(r'[（(][^（）()]*[）)]','',s); return re.sub(r'[〔〕\s　「」]','',s)
def _mfrags(nm):
    return [p for p in re.split(r'[・／/]', _mnorm(nm)) if p]
def _mbest(text, auth):
    t=_mnorm(text); best,bs=0,-1.0
    for i,nm in enumerate(auth):
        sc=0.0
        for fr in _mfrags(nm):
            if t and fr and (t in fr or fr in t): sc=max(sc,0.9+0.1*min(len(t),len(fr))/max(len(t),len(fr)))
            sc=max(sc,difflib.SequenceMatcher(None,t,fr).ratio())
            m=difflib.SequenceMatcher(None,t,fr).find_longest_match(0,len(t),0,len(fr))
            if fr: sc=max(sc,0.6*m.size/len(fr))
        if sc>bs: bs,best=sc,i
    return best,bs
def remap_mi(imi, old_mi, old_meanings, auth):
    # imi（活用形の解答）と旧・抽出ラベルの両方を意味一覧に照合し、最も近い番号(1始まり)を返す。
    # 対応が無い例文(old_mi=0/imi空)は 0（意味番号なし＝「・」表示）を維持。
    if not old_mi or not (imi and imi.strip()): return 0
    if len(auth)==1: return 1
    cands=[_mbest(imi,auth)]
    if 1<=old_mi<=len(old_meanings): cands.append(_mbest(old_meanings[old_mi-1],auth))
    return max(cands,key=lambda c:c[1])[0]+1

# ---------- assemble base ----------
WORD={}; base=[]
for no,sub in df.groupby('no'):
    word=cl_txt(sub['word'].iloc[0]); WORD[no]=word
    pos=cl_txt(sub['pos'].iloc[0]); is_k=sub['keigo'].notna().any()
    meanings=[]; ex=[]
    for _,r in sub.iterrows():
        imi=cl_imi(r['imi'])
        if imi and imi not in meanings: meanings.append(imi)
        ex.append({'imi':imi,'ko':strip_src(r['ko']),'koBlank':strip_src(r['koBlank']),
                   'koU':ko_underline(r['ko'],r['koBlank']),
                   'yk':cl_yk(r['yk']),'ykBlank':cl_yk(r['ykBlank']),'src':src_of(r['ko'])})
    clusters=cluster_meanings(meanings)
    raw_labels=[canon_label(c) for c in clusters]
    llabels=[lemmatize(l) for l in raw_labels]
    final=[]; cl2final=[]
    for l in llabels:
        if l in final: cl2final.append(final.index(l))
        else: final.append(l); cl2final.append(len(final)-1)
    imi2fi={}
    for i,c in enumerate(clusters):
        for m in c: imi2fi[m]=cl2final[i]+1
    for e in ex: e['mi']=imi2fi.get(e['imi'],0)
    base.append({'no':no,'word':word,'pos':pos,'imp':importance(no),
                 'keigo':KEIGO.get(no,[]),'meanings':final,'examples':ex})

# ---------- override meanings with authoritative 意味一覧, remap example mi ----------
_miss=[w['no'] for w in base if w['no'] not in AUTH]
if _miss: print('WARN: 意味一覧に無い番号:', _miss)
for w in base:
    auth=AUTH.get(w['no'])
    if not auth: continue
    old_means=w['meanings']
    for e in w['examples']:
        e['mi']=remap_mi(e['imi'], e['mi'], old_means, auth)
    w['meanings']=list(auth)

# ---------- relations (類義/対義/混同), per-sense ----------
def S(*pairs): return list(pairs)
SYN=[
 S((3,'がまんする'),(5,'がまんする')),
 S((24,'つらい'),(25,'つらい'),(221,'つらい・気の毒'),(233,'つらい')),
 S((12,'かわいい'),(13,'いとしい'),(86,'いとしい・気の毒'),(253,'かわいらしい')),
 S((20,'気がかり'),(85,'待ち遠しい・不安'),(220,'心配'),(291,'気が晴れない')),
 S((14,'すばらしい'),(21,'立派'),(22,'すばらしい'),(88,'すばらしい'),(74,'すばらしい')),
 S((15,'趣がある'),(26,'しみじみ'),(74,'興趣がある')),
 S((7,'結婚する'),(8,'結婚する'),(189,'男のもとに通う')),
 S((66,'訪れる・見舞う'),(186,'訪ねる・手紙を出す')),
 S((69,'患う・思い悩む'),(204,'病気で苦しむ'),(197,'疲れる'),(201,'病気になる')),
 S((72,'亡くなる'),(205,'先立たれる')),
 S((99,'高貴だ'),(76,'高貴だ'),(16,'身分が高い')),
 S((213,'優美だ'),(254,'優美だ'),(288,'気品がある'),(210,'奥ゆかしい'),(100,'美しい')),
 S((29,'退屈・もの寂しい'),(97,'物足りない')),
 S((2,'評判になる'),(147,'うわさに聞く'),(114,'評判'),(60,'世間に知られる')),
 S((127,'すぐに'),(128,'早く'),(129,'早く'),(45,'すぐに')),
 S((62,'準備する'),(206,'用意する'),(305,'準備')),
 S((42,'たくさん'),(269,'たくさん')),
 S((251,'ひそかに'),(274,'そっと')),
 S((4,'思われる'),(7,'思われる')),
]
CONF=[
 S((1,'目を覚ます'),(87,'驚きあきれる'),(237,'気にくわない')),
 S((22,'すばらしい'),(55,'愛する・感嘆'),(218,'感じよい'),(237,'気にくわない')),
 S((155,'申し上げる(謙)'),(168,'参上する(謙)'),(171,'退出する(謙)'),(172,'退出する(謙)'),(169,'参上/差し上げる'),(170,'差し上げる(謙)')),
 S((159,'お与えになる(尊・四段)'),(160,'〜ております(丁・下二)')),
 S((16,'よし＝形容詞'),(122,'よし＝名詞')),
 S((110,'しるし＝名詞(効果)'),(292,'しるし＝形容詞(明白)')),
 S((60,'聞こゆ＝動詞(聞こえる)'),(156,'きこゆ＝敬語(申し上げる)')),
 S((45,'そのまま'),(137,'そのまま・全部'),(316,'そのまま')),
 S((19,'ゆかし＝〜したい'),(77,'ゆゆし＝不吉/立派')),
 S((21,'ありがたし＝めったにない'),(208,'かたじけなし＝おそれ多い')),
 S((89,'つれなし＝冷淡'),(29,'つれづれ＝退屈')),
 S((90,'あいなし＝つまらない'),(91,'あぢきなし＝どうにもならない'),(95,'よしなし＝つまらない')),
 S((69,'患う'),(204,'病気'),(202,'静養する'),(203,'病が癒える'),(240,'病が重い')),
]
ANT=[((99,'高貴だ'),(17,'身分が低い')),((76,'高貴だ'),(17,'身分が低い')),
 ((16,'身分が高い'),(17,'身分が低い・粗末')),((21,'めったにない'),(325,'ありきたりだ')),
 ((21,'めったにない'),(130,'総じて・普通')),((117,'早朝'),(275,'一晩中'))]

rel={no:[] for no in WORD}
def add_clusters(cls,typ):
    for cl in cls:
        for a,sa in cl:
            for b,sb in cl:
                if a==b or b not in WORD:continue
                rel[a].append({'type':typ,'sense':sa,'no':b,'word':WORD[b],'gloss':sb})
def add_ant(pairs):
    for (a,sa),(b,sb) in pairs:
        if a in WORD and b in WORD:
            rel[a].append({'type':'対義','sense':sa,'no':b,'word':WORD[b],'gloss':sb})
            rel[b].append({'type':'対義','sense':sb,'no':a,'word':WORD[a],'gloss':sa})
add_clusters(SYN,'類義'); add_clusters(CONF,'混同'); add_ant(ANT)
# dedup by (type,no) keeping first
for no in rel:
    seen=set(); out=[]
    for r in rel[no]:
        k=(r['type'],r['no'])
        if k in seen:continue
        seen.add(k); out.append(r)
    rel[no]=out

# ---------- finalize words ----------
FL={'GAPB':'現代語とギャップ大','GAPS':'現代語と少しズレ','POLY':'多義語','CORE':'コアで覚える',
    'VAGUE':'漠然系','POLAR':'プラス/マイナス両義','KOOU':'呼応の副詞','NASHI':'「〜なし」型',
    'KANGO':'漢語サ変','KEIGOCARE':'敬語要注意'}
FLAG_ORDER=['GAPB','GAPS','POLY','CORE','VAGUE','POLAR','KOOU','NASHI','KANGO','KEIGOCARE']
words=[]
for w in base:
    no=w['no']; flags=[]; codes=TH.get(no,[])
    if no in GAP_BIG: flags.append(FL['GAPB'])
    elif no in GAP_SMALL: flags.append(FL['GAPS'])
    if len(w['meanings'])>=4: flags.append(FL['POLY'])
    if no in CORE: flags.append(FL['CORE'])
    if no in VAGUE: flags.append(FL['VAGUE'])
    if 'PLUS' in codes and 'MINUS' in codes: flags.append(FL['POLAR'])  # 両テーマ保有から自動導出（TAG_GUIDE §3-(3)）
    if '～' in w['word'] or no in KOOU_EXTRA: flags.append(FL['KOOU'])
    if w['pos']=='形容詞' and w['word'].endswith('なし'): flags.append(FL['NASHI'])
    if no in KANGO: flags.append(FL['KANGO'])
    if no in KEIGO_CARE: flags.append(FL['KEIGOCARE'])
    w['themes']=[TL[c] for c in codes]
    w['flags']=flags
    w['related']=rel[no]
    w['kat']=KATSUYO.get(no,'ナリ活用' if w['pos']=='形容動詞' else '')
    words.append(w)
words.sort(key=lambda x:x['no'])

# ---------- attach explanations from MD ----------
import os
NOTE_MD=os.path.join(ROOT,'data','explanations.md')
notes={}; kanji={}
if os.path.exists(NOTE_MD):
    npat=re.compile(r'^(\d{1,3})\s+(.+?)\s+[—–-]\s+(.+)$')
    for ln in open(NOTE_MD,encoding='utf-8').read().split('\n'):
        m=npat.match(ln.strip())
        if m:
            no=int(m.group(1)); head=m.group(2).strip(); body=m.group(3).strip()
            notes[no]=body
            km=re.search(r'[（(](.+?)[）)]',head)
            if km: kanji[no]=km.group(1).strip()
for w in words:
    w['note']=notes.get(w['no'],'')
    w['kanji']=kanji.get(w['no'],'')
print('解説付与:',sum(1 for w in words if w['note']),'/',len(words),'| 漢字付与:',sum(1 for w in words if w['kanji']))

meta={'imp':['最必修','必修','必修敬語','重要','応用'],
 'pos':['動詞','形容詞','形容動詞','名詞','副詞','連体詞・連語'],
 'keigo':['尊敬','謙譲','丁寧'],
 'theme':[TL[c] for c in THEME_ORDER],
 'flag':[FL[c] for c in FLAG_ORDER],'total':len(words)}

json.dump({'words':words,'meta':meta},open(os.path.join(ROOT,'build','data_payload.json'),'w',encoding='utf-8'),ensure_ascii=False)
print('words',len(words),'| theme',sum(1 for w in words if w['themes']),
      '| flag',sum(1 for w in words if w['flags']),
      '| related',sum(1 for w in words if w['related']))
from collections import Counter
fc=Counter(f for w in words for f in w['flags'])
for k in meta['flag']:print(' ',k,fc.get(k,0))
